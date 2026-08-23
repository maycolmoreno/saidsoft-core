import datetime
import io
import json
import tempfile
from unittest.mock import patch

from cryptography.fernet import Fernet
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase, override_settings

from apps.activos.models import Cargo, Colaborador, Departamento
from apps.catalogo import crypto
from apps.catalogo.models import ClaveRecuperacionBitLocker, Estacion, Farmacia, Grupo, UnidadNegocio, VersionAgente
from apps.catalogo.services import (
    enviar_actualizacion_agente, enviar_comando, enviar_script, firmar_payload,
    generar_comando_instalacion_meshcentral, obtener_clave_bitlocker_descifrada, resolver_estaciones,
    url_escritorio_remoto_meshcentral, url_grabaciones_meshcentral, url_terminal_remoto_meshcentral,
    validar_destino_unidad_negocio,
)

MESHCENTRAL_CONFIG_TEST = {
    'SERVER_URL': 'https://mesh.test.local',
    'MESH_ID': 'abc123meshid',
    'AGENT_ARCH_ID': 4,
    'INSTALL_FLAGS': 2,
    'VIEWMODE_ESCRITORIO': '11',
    'VIEWMODE_TERMINAL': '12',
}

BITLOCKER_KEY_TEST = Fernet.generate_key().decode()


@override_settings(BITLOCKER_ENCRYPTION_KEY=BITLOCKER_KEY_TEST)
class BitLockerCryptoTests(TestCase):
    def test_cifrar_descifrar_es_reversible(self):
        original = '111111-222222-333333-444444-555555-666666-777777-888888'
        token = crypto.cifrar(original)
        self.assertNotEqual(token, original)  # nunca texto plano
        self.assertEqual(crypto.descifrar(token), original)

    def test_descifrar_token_invalido_lanza_value_error(self):
        with self.assertRaises(ValueError):
            crypto.descifrar('esto-no-es-un-token-fernet-valido')

    def test_clave_cifrada_con_otra_llave_no_se_puede_descifrar(self):
        token = crypto.cifrar('secreto')
        with override_settings(BITLOCKER_ENCRYPTION_KEY=Fernet.generate_key().decode()):
            with self.assertRaises(ValueError):
                crypto.descifrar(token)


@override_settings(BITLOCKER_ENCRYPTION_KEY=BITLOCKER_KEY_TEST)
class ObtenerClaveBitlockerTests(TestCase):
    def setUp(self):
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(
            codigo='ML001', grupo=grupo, unidad_negocio=UnidadNegocio.objects.get(codigo='SG'),
        )
        self.estacion = Estacion.objects.create(codigo='ML001-A', farmacia=farmacia)

    def test_none_si_no_hay_clave_registrada(self):
        self.assertIsNone(obtener_clave_bitlocker_descifrada(self.estacion))

    def test_devuelve_la_clave_en_texto_plano_cuando_existe(self):
        ClaveRecuperacionBitLocker.objects.create(
            estacion=self.estacion, clave_cifrada=crypto.cifrar('111111-222222-333333'),
        )
        self.assertEqual(obtener_clave_bitlocker_descifrada(self.estacion), '111111-222222-333333')


@override_settings(MESHCENTRAL_CONFIG=MESHCENTRAL_CONFIG_TEST)
class MeshCentralServiciosTests(TestCase):
    def setUp(self):
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(codigo='ML001', grupo=grupo, unidad_negocio=UnidadNegocio.objects.get(codigo='SG'))
        self.estacion = Estacion.objects.create(codigo='ML001-A', farmacia=farmacia)

    def test_generar_comando_instalacion_contiene_mesh_id_y_url(self):
        comando = generar_comando_instalacion_meshcentral(self.estacion)
        self.assertIn('mesh.test.local', comando)
        self.assertIn('meshid=abc123meshid', comando)
        self.assertIn('installflags=2', comando)

    def test_generar_comando_usa_curl_no_invoke_webrequest(self):
        # Invoke-WebRequest de PowerShell 5.1 corta la descarga con "error inesperado de
        # envío" contra el TLS autofirmado de MeshCentral, incluso forzando Tls12 y
        # saltando la validación del certificado — probado de verdad instalando en
        # ML006-A y MC001-B (ver PLAN_MODERNIZACION.md). curl.exe (Schannel, no el stack
        # de .NET Framework) sí funciona.
        comando = generar_comando_instalacion_meshcentral(self.estacion)
        self.assertIn('curl.exe', comando)
        self.assertNotIn('Invoke-WebRequest', comando)

    def test_generar_comando_usa_fullinstall(self):
        # Corrido sin argumentos, meshagent.exe nunca completa la instalación real
        # cuando lo lanza un servicio de Windows sin sesión interactiva (Session 0) —
        # se queda corriendo suelto sin registrar el servicio "Mesh Agent". Con
        # "-fullinstall" sí instala de verdad y el proceso termina solo (reproducido
        # y diagnosticado en MC001-C), por eso "-Wait" es correcto acá.
        comando = generar_comando_instalacion_meshcentral(self.estacion)
        self.assertIn('-fullinstall', comando)
        self.assertIn('-Wait', comando)

    def test_generar_comando_nombra_el_agente_como_la_estacion(self):
        # Para que apps.monitoreo.adapters.meshcentral._vincular_por_nombre pueda
        # enlazar el node_id solo, sin copiarlo a mano de la consola de MeshCentral.
        comando = generar_comando_instalacion_meshcentral(self.estacion)
        self.assertIn(f'--agentName={self.estacion.codigo}', comando)

    def test_urls_remotas_none_sin_node_id(self):
        self.assertIsNone(url_escritorio_remoto_meshcentral(self.estacion))
        self.assertIsNone(url_terminal_remoto_meshcentral(self.estacion))
        self.assertIsNone(url_grabaciones_meshcentral(self.estacion))

    def test_urls_remotas_con_node_id(self):
        self.estacion.meshcentral_node_id = 'nodeid123'
        self.estacion.save(update_fields=['meshcentral_node_id'])

        url_escritorio = url_escritorio_remoto_meshcentral(self.estacion)
        url_terminal = url_terminal_remoto_meshcentral(self.estacion)
        url_grabaciones = url_grabaciones_meshcentral(self.estacion)

        self.assertIn('gotonode=nodeid123', url_escritorio)
        self.assertIn('viewmode=11', url_escritorio)
        self.assertIn('gotonode=nodeid123', url_terminal)
        self.assertIn('viewmode=12', url_terminal)
        self.assertIn('gotonode=nodeid123', url_grabaciones)


class MultiTenantAislamientoTests(TestCase):
    """R1: un Grupo (canal TRX) puede estar compartido por farmacias de varias
    unidades de negocio (ver docstring de apps.cumplimiento.services) — estos tests
    prueban que eso nunca se traduce en un despliegue/ejecución cruzando de tenant."""

    def setUp(self):
        self.sg = UnidadNegocio.objects.get(codigo='SG')
        self.mia = UnidadNegocio.objects.get(codigo='MIA')
        self.grupo = Grupo.objects.create(codigo='TRX001')
        self.farmacia_sg = Farmacia.objects.create(codigo='ML001', grupo=self.grupo, unidad_negocio=self.sg)
        self.farmacia_mia = Farmacia.objects.create(codigo='MAM01', grupo=self.grupo, unidad_negocio=self.mia)
        self.estacion_sg = Estacion.objects.create(
            codigo='ML001-A', farmacia=self.farmacia_sg, estado_aprobacion=Estacion.EstadoAprobacion.APROBADA,
        )
        self.estacion_mia = Estacion.objects.create(
            codigo='MAM01-A', farmacia=self.farmacia_mia, estado_aprobacion=Estacion.EstadoAprobacion.APROBADA,
        )

    def test_cadena_no_incluye_estaciones_de_otro_tenant(self):
        resultado = resolver_estaciones('cadena', unidad_negocio=self.sg)
        self.assertEqual(list(resultado), [self.estacion_sg])

    def test_grupo_compartido_solo_aporta_estaciones_del_tenant_pedido(self):
        # El Grupo es el mismo para ambas farmacias — la fuga que hay que evitar es
        # que "grupos" devuelva la estación de MIA cuando se pidió para SG.
        resultado = resolver_estaciones('grupos', unidad_negocio=self.sg, grupos=[self.grupo])
        self.assertEqual(list(resultado), [self.estacion_sg])

    def test_validar_destino_rechaza_farmacia_de_otro_tenant(self):
        with self.assertRaises(ValidationError):
            validar_destino_unidad_negocio(self.sg, farmacias=[self.farmacia_mia])

    def test_validar_destino_rechaza_estacion_de_otro_tenant(self):
        with self.assertRaises(ValidationError):
            validar_destino_unidad_negocio(self.sg, estaciones=[self.estacion_mia])

    def test_validar_destino_acepta_targets_del_mismo_tenant(self):
        validar_destino_unidad_negocio(self.sg, farmacias=[self.farmacia_sg], estaciones=[self.estacion_sg])


class MarcarEstacionesOfflineTaskTests(TestCase):
    """CELERY_TASK_ALWAYS_EAGER=True en desarrollo.py hace que .delay() corra sincrónico
    en el mismo proceso — sirve para probar que la tarea está bien registrada y hace lo
    mismo que el comando manual, sin necesitar Redis ni un worker real."""

    def test_delay_marca_offline_igual_que_el_comando(self):
        from datetime import timedelta

        from django.utils import timezone

        from apps.catalogo.tasks import marcar_estaciones_offline_task

        sg = UnidadNegocio.objects.get(codigo='SG')
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(codigo='ML001', grupo=grupo, unidad_negocio=sg)
        estacion = Estacion.objects.create(
            codigo='ML001-A', farmacia=farmacia, estado_aprobacion=Estacion.EstadoAprobacion.APROBADA,
            estado_conexion=Estacion.EstadoConexion.ONLINE,
            ultimo_heartbeat=timezone.now() - timedelta(minutes=20),
        )

        resultado = marcar_estaciones_offline_task.delay()

        estacion.refresh_from_db()
        self.assertEqual(estacion.estado_conexion, Estacion.EstadoConexion.OFFLINE)
        self.assertIn('1 estación', resultado.get())


class ImportarFarmaciasTests(TestCase):
    def setUp(self):
        self.mia = UnidadNegocio.objects.get(codigo='MIA')
        self.sg = UnidadNegocio.objects.get(codigo='SG')

    def _correr(self, contenido_csv, **opciones):
        salida = io.StringIO()
        with tempfile.NamedTemporaryFile('w', suffix='.csv', delete=False, newline='') as tmp:
            tmp.write(contenido_csv)
            ruta = tmp.name
        call_command('importar_farmacias', ruta, stdout=salida, **opciones)
        return salida.getvalue()

    def test_crea_farmacias_deduciendo_unidad_y_grupo_por_prefijo(self):
        csv_contenido = (
            'Ciudad,Id de sitio,Tipo de Enlace,Backup,NODO\n'
            'Ambato,MAM01,PUNTO NET,NO,trx001\n'
            'Puerto Bolivar,GP005,TELCONET,NO,trx002\n'
            'Loja,7DM02,TELCONET,NO,hub_111_6\n'
        )
        salida = self._correr(csv_contenido)

        # 7DIAS no existe todavía como UnidadNegocio en este test — esa fila queda
        # afuera, así que solo se crean las 2 que sí tienen unidad de negocio válida.
        self.assertIn('2 farmacia(s) creada(s)', salida)
        mam01 = Farmacia.objects.get(codigo='MAM01')
        self.assertEqual(mam01.unidad_negocio, self.mia)
        self.assertEqual(mam01.grupo.codigo, 'TRX001')
        self.assertEqual(mam01.ubicacion, 'Ambato')

        gp005 = Farmacia.objects.get(codigo='GP005')
        self.assertEqual(gp005.unidad_negocio, self.sg)

        # 7DIAS no existe todavía como UnidadNegocio en este test — esa fila debe
        # reportarse como error, no crear un tenant nuevo por accidente.
        self.assertFalse(Farmacia.objects.filter(codigo='7DM02').exists())
        self.assertIn('7DIAS no existe', salida)

    def test_re_correr_sin_actualizar_omite_las_que_ya_existen(self):
        grupo = Grupo.objects.create(codigo='TRX001')
        Farmacia.objects.create(codigo='MAM01', grupo=grupo, unidad_negocio=self.mia, ubicacion='Vieja')
        csv_contenido = 'Ciudad,Id de sitio,NODO\nAmbato,MAM01,trx001\n'

        salida = self._correr(csv_contenido)

        self.assertIn('0 farmacia(s) creada(s)', salida)
        self.assertIn('1 farmacia(s) ya existían, omitida(s): MAM01', salida)
        self.assertEqual(Farmacia.objects.get(codigo='MAM01').ubicacion, 'Vieja')

    def test_actualizar_sobreescribe_ubicacion_y_grupo(self):
        grupo_viejo = Grupo.objects.create(codigo='TRX001')
        Farmacia.objects.create(codigo='MAM01', grupo=grupo_viejo, unidad_negocio=self.mia, ubicacion='Vieja')
        csv_contenido = 'Ciudad,Id de sitio,NODO\nAmbato Centro,MAM01,trx002\n'

        self._correr(csv_contenido, actualizar=True)

        mam01 = Farmacia.objects.get(codigo='MAM01')
        self.assertEqual(mam01.ubicacion, 'Ambato Centro')
        self.assertEqual(mam01.grupo.codigo, 'TRX002')

    def test_dry_run_no_escribe_nada(self):
        csv_contenido = 'Ciudad,Id de sitio,NODO\nAmbato,MAM01,trx001\n'

        salida = self._correr(csv_contenido, dry_run=True)

        self.assertIn('[DRY RUN] 1 farmacia(s) creada(s)', salida)
        self.assertFalse(Farmacia.objects.filter(codigo='MAM01').exists())
        self.assertFalse(Grupo.objects.filter(codigo='TRX001').exists())

    def test_prefijo_desconocido_se_reporta_como_error_sin_adivinar(self):
        csv_contenido = 'Ciudad,Id de sitio,NODO\nQuito,ZQ001,trx001\n'

        salida = self._correr(csv_contenido)

        self.assertFalse(Farmacia.objects.filter(codigo='ZQ001').exists())
        self.assertIn('prefijo de código sin mapeo', salida)

    def test_nodo_mas_largo_que_el_campo_se_reporta_como_error_sin_reventar(self):
        # Grupo.codigo tiene max_length=10 — un NODO más largo tiraba un DataError sin
        # manejar (500 crudo en el admin, encontrado en producción 12-ago-2026).
        csv_contenido = 'Ciudad,Id de,NODO\nAmbato,MAM01,un_nodo_con_nombre_demasiado_largo\n'

        salida = self._correr(csv_contenido)

        self.assertFalse(Farmacia.objects.filter(codigo='MAM01').exists())
        self.assertFalse(Grupo.objects.filter(codigo__startswith='UN_NODO').exists())
        self.assertIn('caracteres (máximo', salida)

    def test_provincia_se_combina_con_ciudad_en_ubicacion(self):
        csv_contenido = 'Provincia,Ciudad,Id de,NODO\nEl Oro,Pasaje,MP001,trx001\n'

        self._correr(csv_contenido)

        self.assertEqual(Farmacia.objects.get(codigo='MP001').ubicacion, 'Pasaje, El Oro')

    def test_captura_segmento_red_tipo_enlace_y_backup(self):
        csv_contenido = (
            'Ciudad,Id de,Segmento de Red,Tipo de Enlace,Login,Backup,NODO\n'
            'Pasaje,MP001,10.110.1.96/27,TELCONET,farmamia-mp001,ACTIVO,trx001\n'
            'Pindal,MPDL1,10.101.22.192/27,TELCONET,farmamia-mpdl1,,trx001\n'
        )

        self._correr(csv_contenido)

        mp001 = Farmacia.objects.get(codigo='MP001')
        self.assertEqual(mp001.segmento_red, '10.110.1.96/27')
        self.assertEqual(mp001.tipo_enlace, 'TELCONET')
        self.assertTrue(mp001.tiene_backup)

        # Backup vacío -> sin enlace de respaldo.
        self.assertFalse(Farmacia.objects.get(codigo='MPDL1').tiene_backup)

    def test_captura_ip_por_farmacia_no_por_nodo(self):
        # La IP es por farmacia, no por nodo/grupo -- un mismo NODO puede agrupar
        # farmacias con IPs distintas (rollout de versión de POS, no topología de red).
        csv_contenido = (
            'Ciudad,Id de,NODO,IP\n'
            'Pasaje,MP001,trx001,192.168.112.5\n'
            'Pinas,MI001,trx001,192.168.112.60\n'
        )
        self._correr(csv_contenido)
        self.assertEqual(Farmacia.objects.get(codigo='MP001').ip_router, '192.168.112.5')
        self.assertEqual(Farmacia.objects.get(codigo='MI001').ip_router, '192.168.112.60')

    def test_ip_invalida_se_reporta_como_error_sin_bloquear_la_fila(self):
        csv_contenido = 'Ciudad,Id de,NODO,IP\nPasaje,MP001,trx001,no-es-una-ip\n'
        salida = self._correr(csv_contenido)
        self.assertIn('IP "no-es-una-ip" inválida', salida)
        mp001 = Farmacia.objects.get(codigo='MP001')
        self.assertIsNone(mp001.ip_router)


class FarmaciaAdminImportarViewTests(TestCase):
    """El botón "Importar CSV" del admin (/admin/catalogo/farmacia/importar/) usa el
    mismo apps.catalogo.services.importar_farmacias_desde_csv que el comando de
    management — surgió porque el usuario buscaba un botón de importar en el admin
    y no había ninguno, solo el comando por SSH."""

    def setUp(self):
        from django.contrib.auth.models import User
        self.mia = UnidadNegocio.objects.get(codigo='MIA')
        self.staff = User.objects.create_user(username='staff_import', password='x', is_staff=True)
        self.staff.user_permissions.add(*self._permisos_farmacia())
        self.sin_permiso = User.objects.create_user(username='sin_permiso_import', password='x', is_staff=True)

    @staticmethod
    def _permisos_farmacia():
        from django.contrib.auth.models import Permission
        return Permission.objects.filter(content_type__app_label='catalogo', content_type__model='farmacia')

    def _url(self):
        from django.urls import reverse
        return reverse('admin:catalogo_farmacia_importar')

    def _archivo(self, contenido):
        return io.BytesIO(contenido.encode('utf-8'))

    def test_get_muestra_el_formulario(self):
        self.client.force_login(self.staff)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Importar farmacias desde CSV')

    def test_sin_permiso_de_alta_da_403(self):
        self.client.force_login(self.sin_permiso)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 403)

    def test_post_con_dry_run_no_escribe_y_muestra_previsualizacion(self):
        self.client.force_login(self.staff)
        archivo = self._archivo('Ciudad,Id de,NODO\nAmbato,MAM01,trx001\n')

        resp = self.client.post(self._url(), {'archivo': archivo, 'dry_run': 'on'})

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Previsualización')
        self.assertFalse(Farmacia.objects.filter(codigo='MAM01').exists())

    def test_post_sin_dry_run_crea_y_redirige_al_listado(self):
        from django.urls import reverse
        self.client.force_login(self.staff)
        archivo = self._archivo('Ciudad,Id de,NODO\nAmbato,MAM01,trx001\n')

        resp = self.client.post(self._url(), {'archivo': archivo})

        self.assertRedirects(resp, reverse('admin:catalogo_farmacia_changelist'))
        self.assertTrue(Farmacia.objects.filter(codigo='MAM01').exists())

    def test_post_sin_archivo_muestra_error(self):
        self.client.force_login(self.staff)

        resp = self.client.post(self._url(), {'dry_run': 'on'})

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Elegí un archivo CSV')


class EstacionAdminMonitoreoEnLoteTests(TestCase):
    """list_editable alcanza fila por fila, pero no escala a ~1.800 estaciones — estas
    dos acciones del admin activan/desactivan monitorear_recursos sobre la selección
    completa (filtrable por grupo/farmacia con list_filter)."""

    def setUp(self):
        from django.contrib.auth.models import User
        self.admin_user = User.objects.create_superuser(username='admin_mon', email='a@a.com', password='x')
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(codigo='ML001', grupo=grupo, unidad_negocio=UnidadNegocio.objects.get(codigo='SG'))
        self.e1 = Estacion.objects.create(codigo='ML001-A', farmacia=farmacia)
        self.e2 = Estacion.objects.create(codigo='ML001-B', farmacia=farmacia, monitorear_recursos=True)
        self.client.force_login(self.admin_user)

    def _url(self):
        from django.urls import reverse
        return reverse('admin:catalogo_estacion_changelist')

    def test_activar_en_lote_solo_toca_las_que_no_lo_tenian(self):
        resp = self.client.post(self._url(), {
            'action': 'activar_monitoreo_recursos', '_selected_action': [self.e1.pk, self.e2.pk],
        }, follow=True)
        self.e1.refresh_from_db()
        self.e2.refresh_from_db()
        self.assertTrue(self.e1.monitorear_recursos)
        self.assertTrue(self.e2.monitorear_recursos)
        self.assertContains(resp, 'activado en 1 estación')  # e2 ya estaba activa, se excluye

    def test_desactivar_en_lote(self):
        resp = self.client.post(self._url(), {
            'action': 'desactivar_monitoreo_recursos', '_selected_action': [self.e1.pk, self.e2.pk],
        }, follow=True)
        self.e2.refresh_from_db()
        self.assertFalse(self.e2.monitorear_recursos)
        self.assertContains(resp, 'desactivado en 1 estación')

    def test_activar_en_lote_audita_cada_estacion(self):
        from apps.auditoria.models import EventoAuditoria
        self.client.post(self._url(), {
            'action': 'activar_monitoreo_recursos', '_selected_action': [self.e1.pk],
        })
        self.assertTrue(
            EventoAuditoria.objects.filter(accion='estacion.monitoreo_activar', usuario=self.admin_user).exists(),
        )


class ImportarRedFarmaciasXlsxTests(TestCase):
    """Wrapper sobre importar_farmacias_desde_csv que lee directo del Excel real de
    red (dos hojas, FARMAMIA y SAN GREGORIO, con columnas distintas entre sí)."""

    def setUp(self):
        self.mia = UnidadNegocio.objects.get(codigo='MIA')
        self.sg = UnidadNegocio.objects.get(codigo='SG')

    def _libro_de_prueba(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        farmamia = wb.create_sheet('FARMAMIA')
        farmamia.append(['mcu', 'Provincia', 'Ciudad', 'Id de Farmacia', 'Segmento de Red', 'Tipo de Enlace', 'Login', 'Backup', 'IP-DNS', 'Correo', 'NODO', 'IP'])
        farmamia.append([1, 'El Oro', 'Arenillas', 'MA001', '10.101.18.224/27', 'TELCONET', 'login1', 'ACTIVO', None, None, 'trx001', '192.168.112.5'])

        sg = wb.create_sheet('SAN GREGORIO')
        sg.append(['Item', 'Provincia', 'Canton', 'Direccion', 'Id de Farmacia', 'Login', 'Backup', 'Proveedor', 'RED LAN', 'NODO', 'IP', 'CLAVE'])
        sg.append([2, 'MANABI', 'SANTA ANA', 'Direccion X', 'GSA01', 'login2', None, 'TELCONET', '192.168.102.1', 'trx003', '192.168.112.60', None])

        ruta = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        wb.save(ruta)
        return ruta

    def _correr(self, **opciones):
        salida = io.StringIO()
        call_command('importar_red_farmacias_xlsx', self._libro_de_prueba(), stdout=salida, **opciones)
        return salida.getvalue()

    def test_crea_farmacias_de_ambas_hojas_con_su_propia_ip(self):
        salida = self._correr()
        self.assertIn('2 farmacia(s) creada(s)', salida)

        ma001 = Farmacia.objects.get(codigo='MA001')
        self.assertEqual(ma001.unidad_negocio, self.mia)
        self.assertEqual(ma001.grupo.codigo, 'TRX001')
        self.assertEqual(ma001.ip_router, '192.168.112.5')
        self.assertEqual(ma001.segmento_red, '10.101.18.224/27')

        gsa01 = Farmacia.objects.get(codigo='GSA01')
        self.assertEqual(gsa01.unidad_negocio, self.sg)
        self.assertEqual(gsa01.grupo.codigo, 'TRX003')
        self.assertEqual(gsa01.ip_router, '192.168.112.60')
        self.assertEqual(gsa01.tipo_enlace, 'TELCONET')

    def test_dry_run_no_escribe_nada(self):
        salida = self._correr(dry_run=True)
        self.assertIn('[DRY RUN] 2 farmacia(s) creada(s)', salida)
        self.assertFalse(Farmacia.objects.filter(codigo='MA001').exists())

    def test_nodo_sin_asignar_se_remapea_a_pendiente(self):
        # "ELIPSYS_CRESIO" es el valor real que trae el NODO de las sucursales sin
        # canal de versión de POS asignado todavía -- no es un grupo real y además
        # excede Grupo.codigo (max_length=10), así que se remapea a un placeholder.
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sg = wb.create_sheet('SAN GREGORIO')
        sg.append(['Item', 'Provincia', 'Canton', 'Direccion', 'Id de Farmacia', 'Login', 'Backup', 'Proveedor', 'RED LAN', 'NODO', 'IP', 'CLAVE'])
        sg.append([1, 'MANABI', 'SANTA ANA', 'Direccion X', 'GSA02', 'login', None, 'TELCONET', '192.168.102.1', 'elipsys_cresio', '192.168.112.61', None])
        ruta = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        wb.save(ruta)

        salida = io.StringIO()
        call_command('importar_red_farmacias_xlsx', ruta, stdout=salida)

        self.assertIn('1 farmacia(s) creada(s)', salida.getvalue())
        self.assertEqual(Farmacia.objects.get(codigo='GSA02').grupo.codigo, 'PENDIENTE')


class ImportarDirectorioSucursalesTests(TestCase):
    """Enriquecimiento de Farmacia ya existentes con el directorio de sucursales de
    RRHH (nombre, horario, coordinadores, coordenadas, técnico asignado, etc.)."""

    def setUp(self):
        self.sg = UnidadNegocio.objects.get(codigo='SG')
        grupo = Grupo.objects.create(codigo='TRX001')
        self.farmacia = Farmacia.objects.create(codigo='GES07', grupo=grupo, unidad_negocio=self.sg)

        departamento = Departamento.objects.create(nombre='Tecnologías e Innovación', tipo=Departamento.Tipo.TECNICO)
        cargo = Cargo.objects.create(nombre='Asistente de Soporte Técnico', departamento=departamento)
        self.tecnico = Colaborador.objects.create(
            nombre='Carranza Cedeño Jaime Leonerys', cedula='1312655291', cargo=cargo,
        )

    def _libro_de_prueba(self, tecnico='JAIME CARRANZA', tipo_sucursal='PROPIA', formato='MOSTRADOR'):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = 'Directorio Personal'
        hoja = wb.active
        hoja.append([
            'Nombre Sucursal', 'Marca', 'Ciudad', 'Sucursal', 'Horario', 'Administrador', 'Coordinador_Zonal',
            'Ext_Ip', 'Celular', 'Correo_Electonico', 'Provincia', 'Coordinador_Regional', 'Direccion',
            'Tipo_Sucursal', 'Latitud', 'Longitud', 'formato_farmacia', 'Parroquia', 'fecha_inicio_op',
            'fecha_inicio_ruc', 'Tecnico',
        ])
        hoja.append([
            'FARMACIAS SAN GREGORIO GES07', 'SAN GREGORIO', 'ESMERALDAS', 'GES07',
            'LUNES A VIERNES: 08:00 - 19:00', 'SASINTUÑA BONE MARTHA', 'MARQUEZ MOSQUERA MARIO',
            None, '0990051735', 'ges07.avlibertad@sangregorio.com.ec', 'ESMERALDAS',
            'ALARCON MACIAS GEOVANNY', 'AVENIDA LIBERTAD / SN', tipo_sucursal, '0.9704555', '-79.6530856',
            formato, 'ESMERALDAS', datetime.datetime(2022, 2, 7), datetime.datetime(2022, 2, 7), tecnico,
        ])
        ruta = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        wb.save(ruta)
        return ruta

    def _correr(self, **kwargs):
        salida = io.StringIO()
        call_command('importar_directorio_sucursales', self._libro_de_prueba(**kwargs), stdout=salida)
        return salida.getvalue()

    def test_enriquece_la_farmacia_existente_con_todos_los_campos(self):
        self._correr()
        self.farmacia.refresh_from_db()
        self.assertEqual(self.farmacia.nombre, 'FARMACIAS SAN GREGORIO GES07')
        self.assertEqual(self.farmacia.administrador, 'SASINTUÑA BONE MARTHA')
        self.assertEqual(self.farmacia.coordinador_zonal, 'MARQUEZ MOSQUERA MARIO')
        self.assertEqual(self.farmacia.coordinador_regional, 'ALARCON MACIAS GEOVANNY')
        self.assertEqual(self.farmacia.ciudad, 'ESMERALDAS')
        self.assertEqual(self.farmacia.provincia, 'ESMERALDAS')
        self.assertEqual(self.farmacia.direccion, 'AVENIDA LIBERTAD / SN')
        self.assertEqual(self.farmacia.tipo_sucursal, Farmacia.TipoSucursal.PROPIA)
        self.assertEqual(self.farmacia.formato_farmacia, Farmacia.FormatoFarmacia.MOSTRADOR)
        self.assertAlmostEqual(self.farmacia.latitud, 0.9704555)
        self.assertAlmostEqual(self.farmacia.longitud, -79.6530856)
        self.assertEqual(self.farmacia.telefono, '0990051735')
        self.assertEqual(self.farmacia.email, 'ges07.avlibertad@sangregorio.com.ec')
        self.assertEqual(self.farmacia.fecha_inicio_operacion, datetime.date(2022, 2, 7))
        self.assertEqual(self.farmacia.tecnico_asignado, self.tecnico)

    def test_formato_mostrador_xp_se_normaliza_con_guion_bajo(self):
        self._correr(formato='MOSTRADOR XP')
        self.farmacia.refresh_from_db()
        self.assertEqual(self.farmacia.formato_farmacia, Farmacia.FormatoFarmacia.MOSTRADOR_XP)

    def test_tecnico_nd_no_vincula_a_nadie(self):
        self._correr(tecnico='N/D')
        self.farmacia.refresh_from_db()
        self.assertIsNone(self.farmacia.tecnico_asignado)

    def test_tecnico_desconocido_se_reporta_como_advertencia(self):
        salida = self._correr(tecnico='ALGUIEN NUEVO')
        self.assertIn('técnico sin mapeo', salida)
        self.farmacia.refresh_from_db()
        self.assertIsNone(self.farmacia.tecnico_asignado)

    def test_codigo_sin_farmacia_todavia_se_reporta_como_error(self):
        self.farmacia.delete()
        salida = self._correr()
        self.assertIn('sin Farmacia todavía en SAIDSOFT', salida)

    def test_dry_run_no_escribe_nada(self):
        self._correr(tipo_sucursal='ASOCIADO')  # deja el estado real limpio primero
        self.farmacia.refresh_from_db()
        self.assertEqual(self.farmacia.tipo_sucursal, Farmacia.TipoSucursal.ASOCIADO)

        salida = io.StringIO()
        call_command(
            'importar_directorio_sucursales', self._libro_de_prueba(tipo_sucursal='PROPIA'),
            dry_run=True, stdout=salida,
        )
        self.assertIn('[DRY RUN] 1 farmacia', salida.getvalue())
        self.farmacia.refresh_from_db()
        self.assertEqual(self.farmacia.tipo_sucursal, Farmacia.TipoSucursal.ASOCIADO)  # sin cambios


class ComandoFirmadoTests(TestCase):
    """SEC-1 (auditoría 22-ago-2026): la firma HMAC de un comando sin parámetros
    (`enviar_comando`) era un string constante ("reiniciar", "consultar_info", ...) —
    la misma firma servía para siempre y para cualquier estación. Ahora `estacion` y
    `timestamp` entran a la firma, y el agente valida ambos (ver agente-prueba/agente_prueba.py)."""

    def setUp(self):
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(codigo='ML001', grupo=grupo, unidad_negocio=UnidadNegocio.objects.get(codigo='SG'))
        self.estacion = Estacion.objects.create(codigo='ML001-A', farmacia=farmacia)
        self.otra_estacion = Estacion.objects.create(codigo='ML001-B', farmacia=farmacia)

    def _publicar(self, estacion, comando):
        with patch('apps.catalogo.services.mqtt_publish.single') as mock_single:
            enviar_comando(estacion, comando)
        return mock_single.call_args.args[1]  # (topico, payload_json, ...)

    def test_el_payload_lleva_estacion_timestamp_y_firma_valida(self):
        payload = json.loads(self._publicar(self.estacion, 'reiniciar'))
        self.assertEqual(payload['estacion'], 'ML001-A')
        self.assertIn('timestamp', payload)
        firma_esperada = firmar_payload(comando='reiniciar', estacion='ML001-A', timestamp=payload['timestamp'])
        self.assertEqual(payload['firma'], firma_esperada)

    def test_la_firma_no_es_constante_entre_invocaciones(self):
        # Antes del fix, firmar_payload(comando='reiniciar') no dependía de nada más:
        # la firma de dos invocaciones cualquiera era exactamente la misma.
        payload_1 = json.loads(self._publicar(self.estacion, 'reiniciar'))
        payload_2 = json.loads(self._publicar(self.estacion, 'reiniciar'))
        # Incluso repitiendo la misma estación, si el timestamp cambia la firma cambia.
        if payload_1['timestamp'] != payload_2['timestamp']:
            self.assertNotEqual(payload_1['firma'], payload_2['firma'])

    def test_la_firma_de_una_estacion_no_sirve_para_otra(self):
        payload = json.loads(self._publicar(self.estacion, 'reiniciar'))
        # Reconstruir la firma que el agente de OTRA estación calcularía (su propio
        # código en vez del que venía en el mensaje) no matchea la que llegó.
        firma_para_otra = firmar_payload(
            comando='reiniciar', estacion=self.otra_estacion.codigo, timestamp=payload['timestamp'],
        )
        self.assertNotEqual(payload['firma'], firma_para_otra)


class ScriptFirmadoTests(TestCase):
    def setUp(self):
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(codigo='ML001', grupo=grupo, unidad_negocio=UnidadNegocio.objects.get(codigo='SG'))
        self.estacion = Estacion.objects.create(codigo='ML001-A', farmacia=farmacia)

    def test_el_payload_lleva_estacion_timestamp_y_firma_valida(self):
        with patch('apps.catalogo.services.mqtt_publish.single') as mock_single:
            enviar_script(
                self.estacion, ejecucion_id=1, resultado_id=2, tipo_script='powershell',
                contenido='Write-Host hola', timeout_segundos=60,
            )
        payload = json.loads(mock_single.call_args.args[1])
        self.assertEqual(payload['estacion'], 'ML001-A')
        self.assertIn('timestamp', payload)
        firma_esperada = firmar_payload(
            comando='ejecutar_script', ejecucion_id=1, resultado_id=2, tipo_script='powershell',
            timeout_segundos=60, contenido='Write-Host hola',
            estacion='ML001-A', timestamp=payload['timestamp'],
        )
        self.assertEqual(payload['firma'], firma_esperada)


class VersionAgenteTests(TestCase):
    """Actualización remota del agente desde el panel: VersionAgente es el equivalente
    de VersionAplicacion (software) pero para el propio binario del agente."""

    def setUp(self):
        self.usuario = User.objects.create_user(username='u', password='x')

    def test_calcula_sha256_y_tamanio_al_guardar(self):
        version = VersionAgente.objects.create(
            version='agente-prueba-0.2', ejecutable=SimpleUploadedFile('agente.exe', b'contenido-falso'),
            creado_por=self.usuario,
        )
        self.assertTrue(version.sha256)
        self.assertEqual(version.tamanio_bytes, len(b'contenido-falso'))

    def test_no_recalcula_el_hash_si_ya_existe(self):
        version = VersionAgente.objects.create(
            version='agente-prueba-0.2', ejecutable=SimpleUploadedFile('agente.exe', b'contenido-falso'),
            creado_por=self.usuario,
        )
        hash_original = version.sha256
        version.notas = 'build de prueba'
        version.save()
        self.assertEqual(version.sha256, hash_original)


class EnviarActualizacionAgenteTests(TestCase):
    def setUp(self):
        grupo = Grupo.objects.create(codigo='TRX001')
        farmacia = Farmacia.objects.create(codigo='ML001', grupo=grupo, unidad_negocio=UnidadNegocio.objects.get(codigo='SG'))
        self.estacion = Estacion.objects.create(codigo='ML001-A', farmacia=farmacia)
        usuario = User.objects.create_user(username='u', password='x')
        self.version = VersionAgente.objects.create(
            version='agente-prueba-0.2', ejecutable=SimpleUploadedFile('agente.exe', b'contenido-falso'),
            creado_por=usuario,
        )

    def test_el_payload_lleva_estacion_timestamp_y_firma_valida(self):
        with patch('apps.catalogo.services.mqtt_publish.single') as mock_single:
            enviar_actualizacion_agente(self.estacion, self.version)
        payload = json.loads(mock_single.call_args.args[1])
        self.assertEqual(payload['comando'], 'actualizar_agente')
        self.assertEqual(payload['estacion'], 'ML001-A')
        self.assertEqual(payload['version'], 'agente-prueba-0.2')
        self.assertEqual(payload['sha256'], self.version.sha256)
        firma_esperada = firmar_payload(
            comando='actualizar_agente', version=payload['version'], url=payload['url'], sha256=payload['sha256'],
            estacion='ML001-A', timestamp=payload['timestamp'],
        )
        self.assertEqual(payload['firma'], firma_esperada)
