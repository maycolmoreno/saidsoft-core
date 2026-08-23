"""Importa el inventario de red de farmacias desde el Excel real de operaciones
("DATOS DE FARMACIAS.xlsx", dos hojas: FARMAMIA y SAN GREGORIO) — wrapper delgado
sobre apps.catalogo.services.importar_farmacias_desde_csv, la misma lógica ya probada
que usa `importar_farmacias` (CSV) y la pantalla de importación del admin.

Normaliza las dos hojas (que traen columnas distintas entre sí — ver detalle abajo) a
un único CSV en memoria con encabezados unificados, y lo pasa tal cual al importador
existente: mismo mapeo de unidad de negocio por prefijo de código (M->MIA, G->SG,
dígito->7DIAS), mismo criterio de "crear el Grupo si el NODO no existe todavía".

Mapeo de columnas por hoja (agosto 2026):
    FARMAMIA:      Id de Farmacia, Ciudad, Provincia, NODO, Segmento de Red,
                   Tipo de Enlace, Backup, IP
    SAN GREGORIO:  Id de Farmacia, Canton (=ciudad), Provincia, NODO, RED LAN
                   (=segmento), Proveedor (=tipo de enlace), Backup, IP

La IP es por FARMACIA, no por NODO/grupo — un mismo nodo puede agrupar farmacias con
IPs distintas (es un canal de versión de POS, no una topología de red), así que nunca
se infiere del grupo.

"ELIPSYS_CRESIO" es el valor real que trae el NODO de las sucursales que todavía no se
asignaron a ningún canal de versión de POS (rollout de TRX001-004/HUB_111_6/HUB1116) —
no es un grupo real, y además excede Grupo.codigo (max_length=10). Se remapea al
placeholder "PENDIENTE", corregible a mano estación por estación cuando cada una entre
en operación real (ver PLAN_MODERNIZACION.md, 22-ago-2026).

Uso:
    python manage.py importar_red_farmacias_xlsx "DATOS DE FARMACIAS(3).xlsx"
    python manage.py importar_red_farmacias_xlsx "DATOS DE FARMACIAS(3).xlsx" --dry-run
    python manage.py importar_red_farmacias_xlsx "DATOS DE FARMACIAS(3).xlsx" --actualizar
"""
import csv
import io

from django.core.management.base import BaseCommand, CommandError

from apps.catalogo.services import importar_farmacias_desde_csv

# (nombre_hoja, columna_codigo, columna_ciudad, columna_provincia, columna_nodo,
#  columna_segmento, columna_tipo_enlace, columna_backup, columna_ip) — 1-indexado,
# tal como vienen las dos hojas reales de "DATOS DE FARMACIAS.xlsx".
HOJAS = [
    ('FARMAMIA', 4, 3, 2, 11, 5, 6, 8, 12),
    ('SAN GREGORIO', 5, 3, 2, 10, 9, 8, 7, 11),
]

NODOS_SIN_ASIGNAR = {'ELIPSYS_CRESIO', '#N/A', ''}
NODO_PLACEHOLDER = 'PENDIENTE'

# Códigos que aparecen en el archivo real de red pero no son farmacias — confirmado con
# el usuario (22-ago-2026). MPREV1 es PREVITAL, la misma entidad que PREV1 en el
# directorio de RRHH, que ya no existe.
CODIGOS_EXCLUIDOS = {'MPREV1'}

ENCABEZADOS = ['codigo', 'ciudad', 'provincia', 'nodo', 'segmento', 'tipo_enlace', 'backup', 'ip']


def _valor_columna(fila, indice_1based):
    valor = fila[indice_1based - 1] if indice_1based - 1 < len(fila) else None
    return '' if valor is None else str(valor).strip()


class Command(BaseCommand):
    help = 'Crea/actualiza farmacias desde el Excel real de red (hojas FARMAMIA + SAN GREGORIO).'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path')
        parser.add_argument('--dry-run', action='store_true', help='No escribe nada, solo muestra qué haría.')
        parser.add_argument(
            '--actualizar', action='store_true',
            help='Si la farmacia ya existe, actualiza red/grupo/ubicación/ip desde el Excel en vez de omitirla.',
        )

    def handle(self, *args, **options):
        import openpyxl

        ruta = options['xlsx_path']
        try:
            libro = openpyxl.load_workbook(ruta, data_only=True)
        except (OSError, KeyError) as exc:
            raise CommandError(f'No se pudo abrir {ruta}: {exc}')

        buffer_csv = io.StringIO()
        escritor = csv.writer(buffer_csv)
        escritor.writerow(ENCABEZADOS)
        total_filas = 0
        for nombre_hoja, c_codigo, c_ciudad, c_provincia, c_nodo, c_segmento, c_enlace, c_backup, c_ip in HOJAS:
            if nombre_hoja not in libro.sheetnames:
                self.stdout.write(self.style.WARNING(f'Hoja "{nombre_hoja}" no encontrada en {ruta}, se omite.'))
                continue
            hoja = libro[nombre_hoja]
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo = _valor_columna(fila, c_codigo).upper()
                if not codigo or codigo in CODIGOS_EXCLUIDOS:
                    continue
                nodo = _valor_columna(fila, c_nodo).upper()
                if nodo in NODOS_SIN_ASIGNAR:
                    nodo = NODO_PLACEHOLDER
                escritor.writerow([
                    codigo, _valor_columna(fila, c_ciudad), _valor_columna(fila, c_provincia),
                    nodo, _valor_columna(fila, c_segmento),
                    _valor_columna(fila, c_enlace), _valor_columna(fila, c_backup), _valor_columna(fila, c_ip),
                ])
                total_filas += 1

        self.stdout.write(f'{total_filas} fila(s) leídas de {len(HOJAS)} hoja(s).')
        buffer_csv.seek(0)

        try:
            resultado = importar_farmacias_desde_csv(
                buffer_csv, dry_run=options['dry_run'], actualizar=options['actualizar'],
            )
        except ValueError as exc:
            raise CommandError(str(exc))

        for nodo in resultado.grupos_nuevos:
            self.stdout.write(f'  + grupo nuevo: {nodo}')

        prefijo = '[DRY RUN] ' if options['dry_run'] else ''
        self.stdout.write(self.style.SUCCESS(
            f'{prefijo}{len(resultado.creadas)} farmacia(s) creada(s): {", ".join(resultado.creadas) or "-"}',
        ))
        if options['actualizar']:
            self.stdout.write(
                f'{prefijo}{len(resultado.actualizadas)} farmacia(s) actualizada(s).',
            )
        else:
            self.stdout.write(
                f'{prefijo}{len(resultado.omitidas)} farmacia(s) ya existían, omitida(s) (usar --actualizar).',
            )
        if resultado.errores:
            self.stdout.write(self.style.ERROR(f'{len(resultado.errores)} fila(s) con error:'))
            for err in resultado.errores:
                self.stdout.write(self.style.ERROR(f'  - {err}'))
