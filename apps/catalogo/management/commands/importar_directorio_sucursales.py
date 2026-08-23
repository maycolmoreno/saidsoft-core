"""Enriquece las Farmacia ya existentes con los datos del directorio de sucursales de
RRHH ("directorio_sucursal_Agosto.xlsx", hoja "Directorio Personal") — nombre
completo, horario, administrador, coordinadores, dirección, tipo/formato de
sucursal, coordenadas, fechas de operación/RUC, teléfono/correo y técnico asignado.

NO crea farmacias — se corre DESPUÉS de `importar_red_farmacias_xlsx` (que sí crea
las que falten, con su Grupo/unidad de negocio correctos). Una fila cuyo código no
existe todavía en SAIDSOFT se reporta como error, nunca se crea a medias sin grupo.

El técnico se vincula por cédula usando el mapeo fijo en
apps.activos.management.commands.crear_tecnicos_soporte.NOMBRE_DIRECTORIO_A_CEDULA
(varias personas figuran en este directorio por su segundo nombre, no alcanza con
comparar texto) — correr `crear_tecnicos_soporte` antes de este comando.

Uso:
    python manage.py importar_directorio_sucursales directorio_sucursal_Agosto.xlsx
    python manage.py importar_directorio_sucursales directorio_sucursal_Agosto.xlsx --dry-run
"""
import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.activos.management.commands.crear_tecnicos_soporte import NOMBRE_DIRECTORIO_A_CEDULA
from apps.activos.models import Colaborador
from apps.catalogo.models import Farmacia

COLUMNAS = {
    'nombre_sucursal': 1, 'marca': 2, 'ciudad': 3, 'sucursal': 4, 'horario': 5,
    'administrador': 6, 'coordinador_zonal': 7, 'extension': 8, 'celular': 9,
    'correo': 10, 'provincia': 11, 'coordinador_regional': 12, 'direccion': 13,
    'tipo_sucursal': 14, 'latitud': 15, 'longitud': 16, 'formato_farmacia': 17,
    'parroquia': 18, 'fecha_inicio_op': 19, 'fecha_inicio_ruc': 20, 'tecnico': 21,
}

CAMPOS_ACTUALIZABLES = [
    'nombre', 'administrador', 'coordinador_zonal', 'coordinador_regional', 'ciudad', 'provincia',
    'parroquia', 'direccion', 'horario', 'tipo_sucursal', 'formato_farmacia', 'latitud', 'longitud',
    'fecha_inicio_operacion', 'fecha_inicio_ruc', 'extension_telefonica', 'telefono', 'email',
    'tecnico_asignado',
]


def _limpiar_texto(valor):
    if valor is None:
        return ''
    # _x000D_ es un artefacto de Excel (retorno de carro escapado en el XML de shared
    # strings) que queda como texto literal en la celda, no un problema real de datos.
    return str(valor).strip().replace('_x000D_', ' ').strip()


def _normalizar_choice(valor):
    return _limpiar_texto(valor).upper().replace(' ', '_').lower()


def _a_fecha(valor):
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    return None


class Command(BaseCommand):
    help = 'Enriquece las Farmacia ya existentes con el directorio de sucursales de RRHH.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path')
        parser.add_argument('--dry-run', action='store_true', help='No escribe nada, solo muestra qué haría.')

    def handle(self, *args, **options):
        import openpyxl

        ruta = options['xlsx_path']
        dry_run = options['dry_run']
        try:
            libro = openpyxl.load_workbook(ruta, data_only=True)
        except (OSError, KeyError) as exc:
            raise CommandError(f'No se pudo abrir {ruta}: {exc}')

        if 'Directorio Personal' not in libro.sheetnames:
            raise CommandError(f'No se encontró la hoja "Directorio Personal" en {ruta}.')
        hoja = libro['Directorio Personal']

        valores_tipo_sucursal = {c[0] for c in Farmacia.TipoSucursal.choices}
        valores_formato = {c[0] for c in Farmacia.FormatoFarmacia.choices}
        colaboradores_por_cedula = {c.cedula: c for c in Colaborador.objects.all()}

        actualizadas, sin_farmacia, sin_tecnico, errores = [], [], [], []

        with transaction.atomic():
            sp = transaction.savepoint()
            for fila_num, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                def _col(clave):
                    indice = COLUMNAS[clave] - 1
                    return fila[indice] if indice < len(fila) else None

                codigo = _limpiar_texto(_col('sucursal')).upper()
                if not codigo:
                    continue

                farmacia = Farmacia.objects.filter(codigo=codigo).first()
                if farmacia is None:
                    sin_farmacia.append(codigo)
                    continue

                tipo_sucursal = _normalizar_choice(_col('tipo_sucursal'))
                if tipo_sucursal and tipo_sucursal not in valores_tipo_sucursal:
                    errores.append(f'fila {fila_num} ({codigo}): tipo_sucursal "{_col("tipo_sucursal")}" desconocido.')
                    tipo_sucursal = ''

                formato_farmacia = _normalizar_choice(_col('formato_farmacia'))
                if formato_farmacia and formato_farmacia not in valores_formato:
                    errores.append(
                        f'fila {fila_num} ({codigo}): formato_farmacia "{_col("formato_farmacia")}" desconocido.',
                    )
                    formato_farmacia = ''

                nombre_tecnico = _limpiar_texto(_col('tecnico')).upper()
                tecnico = None
                if nombre_tecnico and nombre_tecnico != 'N/D':
                    cedula_tecnico = NOMBRE_DIRECTORIO_A_CEDULA.get(nombre_tecnico)
                    tecnico = colaboradores_por_cedula.get(cedula_tecnico) if cedula_tecnico else None
                    if tecnico is None:
                        sin_tecnico.append(f'{codigo} ({nombre_tecnico})')

                extension = _col('extension')
                try:
                    extension = int(extension) if extension not in (None, '') else None
                except (TypeError, ValueError):
                    extension = None

                farmacia.nombre = _limpiar_texto(_col('nombre_sucursal'))
                farmacia.administrador = _limpiar_texto(_col('administrador'))
                farmacia.coordinador_zonal = _limpiar_texto(_col('coordinador_zonal'))
                farmacia.coordinador_regional = _limpiar_texto(_col('coordinador_regional'))
                farmacia.ciudad = _limpiar_texto(_col('ciudad'))
                farmacia.provincia = _limpiar_texto(_col('provincia'))
                farmacia.parroquia = _limpiar_texto(_col('parroquia'))
                farmacia.direccion = _limpiar_texto(_col('direccion'))
                farmacia.horario = _limpiar_texto(_col('horario'))
                farmacia.tipo_sucursal = tipo_sucursal
                farmacia.formato_farmacia = formato_farmacia
                farmacia.latitud = float(_col('latitud')) if _col('latitud') not in (None, '') else None
                farmacia.longitud = float(_col('longitud')) if _col('longitud') not in (None, '') else None
                farmacia.fecha_inicio_operacion = _a_fecha(_col('fecha_inicio_op'))
                farmacia.fecha_inicio_ruc = _a_fecha(_col('fecha_inicio_ruc'))
                farmacia.extension_telefonica = extension
                telefono = _limpiar_texto(_col('celular'))
                if telefono:
                    farmacia.telefono = telefono
                correo = _limpiar_texto(_col('correo'))
                if correo:
                    farmacia.email = correo
                farmacia.tecnico_asignado = tecnico

                if not dry_run:
                    farmacia.save(update_fields=CAMPOS_ACTUALIZABLES)
                actualizadas.append(codigo)

            if dry_run:
                transaction.savepoint_rollback(sp)
            else:
                transaction.savepoint_commit(sp)

        prefijo = '[DRY RUN] ' if dry_run else ''
        self.stdout.write(self.style.SUCCESS(f'{prefijo}{len(actualizadas)} farmacia(s) enriquecida(s).'))
        if sin_farmacia:
            self.stdout.write(self.style.ERROR(
                f'{len(sin_farmacia)} código(s) del directorio sin Farmacia todavía en SAIDSOFT '
                f'(correr importar_red_farmacias_xlsx primero): {", ".join(sin_farmacia)}',
            ))
        if sin_tecnico:
            self.stdout.write(self.style.WARNING(
                f'{len(sin_tecnico)} farmacia(s) con técnico sin mapeo a Colaborador conocido: '
                f'{", ".join(sin_tecnico)}',
            ))
        if errores:
            self.stdout.write(self.style.ERROR(f'{len(errores)} advertencia(s) de datos:'))
            for err in errores:
                self.stdout.write(self.style.ERROR(f'  - {err}'))
